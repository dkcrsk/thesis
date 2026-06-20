"""
data_loader.py
--------------
Reads the three Combibrug Excel files and returns two cleaned DataFrames:
    - employees_df: one row per active employee
    - locations_df: one row per project location with weekly demand

This module contains the data-preparation pipeline described in Section 4
(Data Understanding) of the thesis. The MILP and the UI never touch the
raw Excel files directly — they only consume the cleaned DataFrames.
"""

import re
import pandas as pd
from openpyxl import load_workbook


# Role corrections for rows with missing Functie in the raw medewerkers file.
FUNCTIE_FIXES = {
    31: "Buurtsportcoach",
    51: "Manager",
    57: "Jongerenbegeleider",
    64: "Combicoach",
    88: "Stagiair",
    104: "Stagiair",
    105: "Stagiair",
    107: "Stagiair",
    108: "Stagiair",
    115: "Stagiair",
    117: "Jongerenbegeleider",
    122: "Stagiair",
}

ALL_ROLES = [
    "Combicoach", "Stagiair", "Jongerenbegeleider", "Buurtsportcoach",
    "Manager", "Cultuurcoordinator", "Leefstijlcoach", "Other",
]


class DataLoadError(Exception):
    """Raised when an input file cannot be parsed.

    Use a clear, planner-friendly message — these are surfaced directly
    in the Streamlit UI.
    """
    pass


def _check_required_columns(df, required, file_label):
    """Raise DataLoadError if any required columns are missing."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLoadError(
            f"The {file_label} file is missing required columns: {missing}. "
            f"Columns found: {list(df.columns)[:10]}{'...' if len(df.columns) > 10 else ''}"
        )


def _map_role(raw):
    """Map a raw Functie string to one of the eight clean role categories."""
    if pd.isna(raw):
        return "Other"
    s = str(raw).lower()
    if "stagiair" in s or "stage" in s:
        return "Stagiair"
    if "combicoach" in s or "combi. funct." in s or "combi funct" in s:
        return "Combicoach"
    if "jongerenbegeleider" in s:
        return "Jongerenbegeleider"
    if "buurtsp" in s or "buurtsport" in s:
        return "Buurtsportcoach"
    if "manager" in s or "dga" in s or "locatie leider" in s or "proj.l" in s:
        return "Manager"
    if "cultuur" in s:
        return "Cultuurcoordinator"
    if "leefstijl" in s:
        return "Leefstijlcoach"
    return "Other"


def _map_company(bedrijf):
    if pd.isna(bedrijf):
        return None
    return "CC" if "combicoach" in str(bedrijf).lower() else "Combibrug"


def _map_contract_type(row):
    afd = str(row.get("Afdeling", "")).lower()
    duur = str(row.get("Contractduur", "")).lower()
    if "stage" in afd or "bbl" in afd:
        return "stage"
    if "oproep" in afd:
        return "oproep"
    if "onbepaald" in duur:
        return "permanent"
    return "fixed_term"


def load_employees(employees_path, planning_date="2026-01-01"):
    """Load the medewerkers file and return a cleaned employees DataFrame."""
    try:
        raw = pd.read_excel(employees_path)
    except FileNotFoundError:
        raise DataLoadError(f"Employees file not found: {employees_path}")
    except Exception as e:
        raise DataLoadError(f"Could not open the employees file: {e}")

    _check_required_columns(
        raw,
        required=["Nr", "Functie", "Bedrijf", "Afdeling",
                  "Contractduur", "Contracturen",
                  "Datum uitdienst (laatste dag)"],
        file_label="employees",
    )

    df = raw.copy()
    df["Functie_raw"] = df["Functie"]

    for excel_row, role in FUNCTIE_FIXES.items():
        iloc_idx = excel_row - 2
        if 0 <= iloc_idx < len(df):
            df.iloc[iloc_idx, df.columns.get_loc("Functie")] = role

    df["role"] = df["Functie"].apply(_map_role)
    df["company"] = df["Bedrijf"].apply(_map_company)
    df["contract_type"] = df.apply(_map_contract_type, axis=1)

    planning_date = pd.Timestamp(planning_date)
    end_col = "Datum uitdienst (laatste dag)"
    df["active"] = df[end_col].isna() | (
        pd.to_datetime(df[end_col], errors="coerce") >= planning_date
    )

    employees = df[df["active"]].copy()
    employees = employees.dropna(subset=["Contracturen"])
    employees = employees.rename(columns={
        "Nr": "id",
        "Contracturen": "contract_hours",
    })[["id", "role", "company", "contract_hours", "contract_type"]]

    employees["id"] = employees["id"].astype(int)
    employees["contract_hours"] = employees["contract_hours"].astype(float)
    employees = employees[employees["contract_hours"] > 0]
    employees = employees.drop_duplicates(subset=["id"]).reset_index(drop=True)

    return employees


DEMAND_TEMPLATE_COLUMNS = {
    "site":           "Name of the specific site/location where work happens",
    "company":        "Which company runs it: 'Combibrug' or 'CC'",
    "weekly_hours":   "Expected hours of work needed per week (decimal)",
    "project":            "Project the site belongs to (default = same as site name)",
    "duration_months":    "How many months the project runs (default = full horizon)",
    "start_week":         "Which week of the horizon the project starts in, 1-indexed (default 1)",
    "min_headcount":      "Minimum number of distinct people required per week (default 0, i.e. no requirement). Use 2 for Combiworld/MDT (Four-Eyes rule).",
    "project_leader_id":  "Employee ID of a designated project leader who must work at this site every active week (default: no fixed leader).",
    "available_mon":      "Time windows the site is available on Monday, e.g. '08:00-16:00' or '08:00-12:00; 13:00-17:00'. Blank = closed.",
    "available_tue":      "Time windows the site is available on Tuesday.",
    "available_wed":      "Time windows the site is available on Wednesday.",
    "available_thu":      "Time windows the site is available on Thursday.",
    "available_fri":      "Time windows the site is available on Friday.",
    "notes":              "Free-text notes — not used by the model",
}

DAY_COLS = ["available_mon", "available_tue", "available_wed",
            "available_thu", "available_fri"]
DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri"]


def _parse_windows(cell):
    """Parse a day-availability cell into a list of (start_h, end_h) tuples
    in decimal hours.

    Examples:
        "08:00-16:00"               -> [(8.0, 16.0)]
        "08:00-12:00; 13:00-17:00"  -> [(8.0, 12.0), (13.0, 17.0)]
        "" or NaN                   -> []

    The parser tolerates spaces, en-dashes vs hyphens, and both ':' and '.'
    as the H/M separator. Invalid pieces are skipped silently.
    """
    if cell is None:
        return []
    s = str(cell).strip()
    if s == "" or s.lower() in ("nan", "none"):
        return []

    s = s.replace("–", "-").replace("—", "-")
    out = []
    for piece in s.split(";"):
        piece = piece.strip()
        if not piece:
            continue
        if "-" not in piece:
            continue
        try:
            a, b = piece.split("-", 1)
            def _to_h(t):
                t = t.strip().replace(".", ":")
                if ":" not in t:
                    return float(t)
                hh, mm = t.split(":", 1)
                return int(hh) + int(mm) / 60.0
            start_h = _to_h(a)
            end_h   = _to_h(b)
            if end_h > start_h and 0 <= start_h < 24 and 0 < end_h <= 24:
                out.append((start_h, end_h))
        except (ValueError, IndexError):
            continue
    return out

def load_demand_template(path_or_buffer):
    """Load a forward-looking demand template file.

    Accepts either a file path (str / Path) or a file-like object
    (e.g. a Streamlit UploadedFile). Supports both .xlsx and .csv.

    The template has one row per project location and only needs a
    handful of columns — see DEMAND_TEMPLATE_COLUMNS for the full list.
    This replaces the historical wide-format locations file for future
    planning cycles, where actual timesheets don't exist yet.

    Returns a cleaned locations DataFrame ready for the optimisation model.
    so the rest of the app works identically regardless of which format
    was used.
    """
    try:
        name = getattr(path_or_buffer, "name", str(path_or_buffer))
        if str(name).lower().endswith(".csv"):
            df = pd.read_csv(path_or_buffer)
        else:
            df = pd.read_excel(path_or_buffer)
    except Exception as e:
        raise DataLoadError(f"Could not read the demand template: {e}")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    if "site" not in df.columns and "location" in df.columns:
        df = df.rename(columns={"location": "site"})

    required = ["site", "company", "weekly_hours"]
    missing_cols = [c for c in required if c not in df.columns]
    if missing_cols:
        raise DataLoadError(
            f"The demand template is missing required columns: {missing_cols}. "
            f"Columns found: {list(df.columns)}. "
            f"Required: {required} (the old name 'location' is also accepted for 'site')."
        )

    df = df.dropna(subset=["site", "weekly_hours"])
    df["site"] = df["site"].astype(str).str.strip()
    df["weekly_hours"] = pd.to_numeric(df["weekly_hours"], errors="coerce").fillna(0)
    df = df[df["weekly_hours"] > 0].copy()

    if df.empty:
        raise DataLoadError(
            "The demand template has no valid rows. "
            "Check that 'weekly_hours' contains numbers greater than 0."
        )

    if "duration_months" not in df.columns:
        df["duration_months"] = 12
    else:
        df["duration_months"] = pd.to_numeric(
            df["duration_months"], errors="coerce"
        ).fillna(12).astype(int)

    df["duration_weeks"] = (df["duration_months"] * 52.0 / 12.0).round().astype(int)

    if "start_week" not in df.columns:
        df["start_week"] = 1
    else:
        df["start_week"] = (pd.to_numeric(df["start_week"], errors="coerce")
                            .fillna(1).astype(int).clip(lower=1))

    if "project" not in df.columns:
        df["project"] = df["site"]
    else:
        df["project"] = df["project"].astype(str).str.strip()
        df.loc[df["project"].isin(["", "nan", "None"]), "project"] = (
            df.loc[df["project"].isin(["", "nan", "None"]), "site"]
        )

    if "min_headcount" not in df.columns:
        df["min_headcount"] = 0
    else:
        df["min_headcount"] = (pd.to_numeric(df["min_headcount"], errors="coerce")
                               .fillna(0).astype(int).clip(lower=0))

    if "project_leader_id" not in df.columns:
        df["project_leader_id"] = pd.NA
    else:
        df["project_leader_id"] = pd.to_numeric(
            df["project_leader_id"], errors="coerce"
        ).astype("Int64")

    def _normalise_company(v):
        s = str(v).strip().lower()
        if "cc" in s or "combicoach" in s:
            return "CC"
        return "Combibrug"

    df["company"] = df["company"].apply(_normalise_company)

    for r in ALL_ROLES:
        df[f"req_{r}"] = 0

    df["duration_source"] = "template"

    parsed_windows = []
    totals = []
    for _, row in df.iterrows():
        site_windows = []
        for col, day in zip(DAY_COLS, DAY_NAMES):
            if col in df.columns:
                for (s_h, e_h) in _parse_windows(row.get(col)):
                    site_windows.append((day, s_h, e_h))
        parsed_windows.append(site_windows)
        totals.append(round(sum(e - s for (_, s, e) in site_windows), 2))
    df["windows"] = parsed_windows
    df["total_window_hours"] = totals

    out_cols = [
        "site", "project", "company",
        "weekly_hours", "duration_months", "duration_weeks", "start_week",
        "min_headcount", "project_leader_id",
        "duration_source",
        "windows", "total_window_hours",
    ]
    for c in DAY_COLS:
        if c in df.columns:
            out_cols.append(c)
    out_cols += [f"req_{r}" for r in ALL_ROLES]

    return df[out_cols].reset_index(drop=True)


def generate_demand_template_example(path):
    """Write a blank-but-filled example demand template to path.

    Rosa can use this as a starting point for each new planning cycle.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = [
        ("Dukdalf",                "Dukdalf",              "CC",    8, 12, 1, 0, None,
         "", "08:00-16:00", "", "", "",
         "CC site, Tuesdays only"),
        ("Ley",                    "Ley",                  "CC",   16, 12, 1, 0, None,
         "08:00-13:30", "08:00-13:30", "08:00-13:30", "", "",
         "CC site, Mon-Wed mornings"),
        ("Apollo",                 "Apollo",               "CC",    8, 12, 1, 0, None,
         "", "", "08:00-16:30", "", "",
         "CC site, Wednesdays only"),
        ("Olympiaplein",           "Olympiaplein",         "CC",   16, 12, 1, 0, None,
         "09:00-17:00", "09:00-17:00", "", "", "",
         "CC year-round sports location"),
        ("Mercatorplein",          "Mercatorplein",        "CC",   16, 12, 1, 0, None,
         "", "09:00-17:00", "09:00-17:00", "", "",
         "CC year-round sports location"),
        ("Park Frankendael",       "Combiworld-Oost",      "Combibrug", 12, 10, 1, 2, None,
         "09:00-15:00", "09:00-15:00", "09:00-15:00", "", "",
         "Combiworld — Four-Eyes (min 2 people)"),
        ("Sportpark Drieburg",     "Combiworld-Oost",      "Combibrug",  8, 10, 1, 2, None,
         "", "13:00-17:00", "13:00-17:00", "", "",
         "Combiworld — Four-Eyes"),
        ("Bouwkeet Indische Buurt","Combiworld-Oost",      "Combibrug",  6, 10, 1, 2, None,
         "", "", "", "13:00-17:00", "13:00-17:00",
         "Combiworld — Four-Eyes"),
        ("Buurtcentrum Watergraafsmeer","MDT-Watergraafsmeer","Combibrug",  6,  4, 1, 2, None,
         "09:00-13:00", "13:00-17:00", "", "", "",
         "MDT subsidy, 4 months, Four-Eyes"),
        ("Buurtcentrum Diemen",    "MDT-Diemen",           "Combibrug",   5,  4, 5, 2, None,
         "09:00-13:00", "", "13:00-17:00", "", "",
         "MDT subsidy, starts later"),
        ("BSC Oost",               "BSC-Oost",             "Combibrug",   8, 10, 1, 0, None,
         "09:00-13:00", "", "09:00-13:00", "", "",
         "BSC — fill in project_leader_id with the lead's employee ID"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demand template"

    headers = ["site", "project", "company", "weekly_hours",
               "duration_months", "start_week",
               "min_headcount", "project_leader_id",
               "available_mon", "available_tue", "available_wed",
               "available_thu", "available_fri",
               "notes"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    instructions = [
        "Physical site / location name",
        "Project name (leave blank if same as site)",
        "'Combibrug' or 'CC'",
        "Hours/week (decimal)",
        "Months project runs",
        "Start week (1 = first week)",
        "Min people/week (2 for Combiworld/MDT, 0 otherwise)",
        "Employee ID of fixed project leader (or blank)",
        "Mon time windows, e.g. '08:00-16:00' or '08:00-12:00; 13:00-17:00'",
        "Tue time windows (blank = closed)",
        "Wed time windows",
        "Thu time windows",
        "Fri time windows",
        "Any notes (not used by model)",
    ]
    inst_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                            fill_type="solid")
    for col_idx, inst in enumerate(instructions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=inst)
        cell.fill = inst_fill
        cell.font = Font(italic=True, color="1F4E79")

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx,
                    value="" if val is None else val)

    widths = [28, 22, 12, 14, 17, 14, 16, 19, 18, 18, 18, 18, 18, 45]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = w

    wb.save(path)


EMPLOYEE_TEMPLATE_COLUMNS = {
    "id":             "Employee number (from the staff system)",
    "role":           "Job role (Combicoach, Stagiair, Manager, ...)",
    "company":        "Combibrug or CC",
    "contract_hours": "Contracted hours per week",
    "hourly_cost":    "Cost to Combibrug per hour, € (including employer overhead)",
    "contract_type":  "permanent / fixed_term / stage / oproep (informational)",
    "end_week":       "Last week the employee is available (blank = whole horizon)",
    "is_dreammaker":  "TRUE if this employee is a Dreammaker (Combiworld/MDT), else blank",
}


def generate_employee_template(medewerkers_path, out_path,
                               planning_date="2026-01-01",
                               default_rate=22.0):
    """Create a pre-filled employee template from the raw medewerkers file.

    Rosa runs this once (typically via the Upload tab's helper), then edits
    the resulting Excel file: she fills in the real hourly_cost per person,
    marks any Dreammakers, and sets end_week for anyone leaving mid-horizon.
    After that the planner only ever reads the edited template — the raw
    medewerkers file is no longer needed.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    emp = load_employees(medewerkers_path, planning_date=planning_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = list(EMPLOYEE_TEMPLATE_COLUMNS.keys())
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    inst_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                            fill_type="solid")
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=ci, value=EMPLOYEE_TEMPLATE_COLUMNS[h])
        c.fill = inst_fill
        c.font = Font(italic=True, color="1F4E79")

    for ri, (_, e) in enumerate(emp.iterrows(), start=3):
        ws.cell(row=ri, column=1, value=int(e["id"]))
        ws.cell(row=ri, column=2, value=e["role"])
        ws.cell(row=ri, column=3, value=e["company"])
        ws.cell(row=ri, column=4, value=float(e["contract_hours"]))
        ws.cell(row=ri, column=5, value=default_rate)      # hourly_cost — Rosa edits
        ws.cell(row=ri, column=6, value=e["contract_type"])
        ws.cell(row=ri, column=7, value="")                # end_week — blank
        ws.cell(row=ri, column=8, value="")                # is_dreammaker — blank

    widths = [10, 18, 12, 16, 30, 16, 12, 16]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    wb.save(out_path)


def load_employee_template(path_or_buffer):
    """Load the edited employee template (one combined file with rates)."""
    try:
        df = pd.read_excel(path_or_buffer)
    except Exception as e:
        raise DataLoadError(f"Could not open the employee template: {e}")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    df = df[pd.to_numeric(df["id"], errors="coerce").notna()].copy()

    required = ["id", "role", "company", "contract_hours", "hourly_cost"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLoadError(
            f"The employee template is missing required columns: {missing}. "
            f"Found: {list(df.columns)}"
        )

    df["id"] = pd.to_numeric(df["id"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["id"])
    df["id"] = df["id"].astype(int)
    df["contract_hours"] = pd.to_numeric(df["contract_hours"], errors="coerce").fillna(0).astype(float)
    df["hourly_cost"] = pd.to_numeric(df["hourly_cost"], errors="coerce")

    df["company"] = df["company"].apply(
        lambda v: "CC" if ("cc" in str(v).lower() or "combicoach" in str(v).lower()) else "Combibrug"
    )

    if "contract_type" not in df.columns:
        df["contract_type"] = "unknown"

    # end_week: optional last available week
    if "end_week" not in df.columns:
        df["end_week"] = pd.NA
    else:
        df["end_week"] = pd.to_numeric(df["end_week"], errors="coerce").astype("Int64")

    # is_dreammaker: optional boolean
    if "is_dreammaker" not in df.columns:
        df["is_dreammaker"] = False
    else:
        df["is_dreammaker"] = df["is_dreammaker"].apply(
            lambda v: str(v).strip().lower() in ("true", "1", "yes", "ja", "x")
        )

    df = df[df["contract_hours"] > 0].drop_duplicates(subset=["id"]).reset_index(drop=True)

    missing_rate = df["hourly_cost"].isna()
    df["cost_source"] = "template"
    df.loc[missing_rate, "cost_source"] = "missing"

    out_cols = ["id", "role", "company", "contract_hours", "hourly_cost",
                "contract_type", "end_week", "is_dreammaker", "cost_source"]
    return df[out_cols]
