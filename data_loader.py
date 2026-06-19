"""
data_loader.py
--------------
Reads the three Combibrug Excel files and returns two cleaned DataFrames:
    - employees_df: one row per active employee
    - locations_df: one row per project location with weekly demand

This module contains the data-preparation pipeline described in Section 4
(Data Understanding) of the thesis. The MILP and the UI never touch the
raw Excel files directly — they only consume the cleaned DataFrames.
"""

import re
import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Constants and lookup tables
# ---------------------------------------------------------------------------

# Functie corrections supplied by Rosa van der Vlugt (rows where the role
# was missing in the raw file). Keys are Excel row numbers (1-indexed,
# including the header row), so iloc index = row - 2.
FUNCTIE_FIXES = {
    31: "Buurtsportcoach",
    51: "Manager",
    57: "Jongerenbegeleider",
    64: "Combicoach",
    88: "Stagiair",
    104: "Stagiair",
    105: "Stagiair",
    107: "Stagiair",
    108: "Stagiair",
    115: "Stagiair",
    117: "Jongerenbegeleider",
    122: "Stagiair",
}

ALL_ROLES = [
    "Combicoach", "Stagiair", "Jongerenbegeleider", "Buurtsportcoach",
    "Manager", "Cultuurcoordinator", "Leefstijlcoach", "Other",
]

# Colour-to-duration mapping from Rosa's legend in the locations file.
# Orange  = MDT subsidy projects (~4 months)
# Yellow  = school-year projects (~10 months)
COLOR_TO_DURATION = {
    "FFFFC000": 4,
    "FFFFFF00": 10,
}

WEEKS_IN_FILE = 52  # the historical locations file covers ~1 year


# ---------------------------------------------------------------------------
# Custom exception so the UI can show clean messages
# ---------------------------------------------------------------------------

class DataLoadError(Exception):
    """Raised when an input file cannot be parsed.

    Use a clear, planner-friendly message — these are surfaced directly
    in the Streamlit UI.
    """
    pass


def _check_required_columns(df, required, file_label):
    """Raise DataLoadError if any required columns are missing."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLoadError(
            f"The {file_label} file is missing required columns: {missing}. "
            f"Columns found: {list(df.columns)[:10]}{'...' if len(df.columns) > 10 else ''}"
        )


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def hhmm_to_hours(s):
    """Convert an 'HH:MM' string to decimal hours."""
    if pd.isna(s) or s in ("", "00:00", 0):
        return 0.0
    s = str(s).strip()
    if ":" not in s:
        try:
            return float(s)
        except ValueError:
            return 0.0
    try:
        h, m = s.split(":")
        return int(h) + int(m) / 60.0
    except (ValueError, IndexError):
        return 0.0


def _map_role(raw):
    """Map a raw Functie string to one of the eight clean role categories."""
    if pd.isna(raw):
        return "Other"
    s = str(raw).lower()
    if "stagiair" in s or "stage" in s:
        return "Stagiair"
    if "combicoach" in s or "combi. funct." in s or "combi funct" in s:
        return "Combicoach"
    if "jongerenbegeleider" in s:
        return "Jongerenbegeleider"
    if "buurtsp" in s or "buurtsport" in s:
        return "Buurtsportcoach"
    if "manager" in s or "dga" in s or "locatie leider" in s or "proj.l" in s:
        return "Manager"
    if "cultuur" in s:
        return "Cultuurcoordinator"
    if "leefstijl" in s:
        return "Leefstijlcoach"
    return "Other"


def _map_company(bedrijf):
    if pd.isna(bedrijf):
        return None
    return "CC" if "combicoach" in str(bedrijf).lower() else "Combibrug"


def _map_contract_type(row):
    afd = str(row.get("Afdeling", "")).lower()
    duur = str(row.get("Contractduur", "")).lower()
    if "stage" in afd or "bbl" in afd:
        return "stage"
    if "oproep" in afd:
        return "oproep"
    if "onbepaald" in duur:
        return "permanent"
    return "fixed_term"


def _parse_duration_text(t):
    """Parse a duration cell into (months, source_tag).

    Accepted formats:
      - explicit number of months: "3 maanden", "6 maanden", "1 maand"
      - explicit number of weeks:  "2 weken" (converted to months)
      - "niet van toepassing" / "n.v.t."  -> 0 (treated as no duration)

    Returns (months, source_tag) or (None, None) if nothing matched.
    The cell-colour fallback (orange / yellow) is applied separately by
    load_locations() when this parser returns None.

    Source tags are used for transparency in the diagnostics panel.
    """
    if t is None or (isinstance(t, float) and pd.isna(t)):
        return None, None
    s = str(t).lower().strip()
    if not s:
        return None, None

    # Explicit numeric forms — months win over weeks.
    m = re.search(r"(\d+)\s*maand", s)
    if m:
        return int(m.group(1)), "text:months"
    m = re.search(r"(\d+)\s*we[ek]+", s)
    if m:
        return max(1, int(m.group(1)) // 4), "text:weeks"

    # Explicit "not applicable" marker.
    if "niet van toepassing" in s or "n.v.t" in s:
        return 0, "text:nvt"

    return None, None


# ---------------------------------------------------------------------------
# Step 1 — load and clean the employee file
# ---------------------------------------------------------------------------

def load_employees(employees_path, planning_date="2026-01-01"):
    """Load the medewerkers file and return a cleaned employees DataFrame."""
    try:
        raw = pd.read_excel(employees_path)
    except FileNotFoundError:
        raise DataLoadError(f"Employees file not found: {employees_path}")
    except Exception as e:
        raise DataLoadError(f"Could not open the employees file: {e}")

    _check_required_columns(
        raw,
        required=["Nr", "Functie", "Bedrijf", "Afdeling",
                  "Contractduur", "Contracturen",
                  "Datum uitdienst (laatste dag)"],
        file_label="employees",
    )

    df = raw.copy()
    df["Functie_raw"] = df["Functie"]

    # Apply Rosa's row-level corrections to missing Functie values
    for excel_row, role in FUNCTIE_FIXES.items():
        iloc_idx = excel_row - 2
        if 0 <= iloc_idx < len(df):
            df.iloc[iloc_idx, df.columns.get_loc("Functie")] = role

    df["role"] = df["Functie"].apply(_map_role)
    df["company"] = df["Bedrijf"].apply(_map_company)
    df["contract_type"] = df.apply(_map_contract_type, axis=1)

    # Active = no end date, or end date >= planning date
    planning_date = pd.Timestamp(planning_date)
    end_col = "Datum uitdienst (laatste dag)"
    df["active"] = df[end_col].isna() | (
        pd.to_datetime(df[end_col], errors="coerce") >= planning_date
    )

    employees = df[df["active"]].copy()
    employees = employees.dropna(subset=["Contracturen"])
    employees = employees.rename(columns={
        "Nr": "id",
        "Contracturen": "contract_hours",
    })[["id", "role", "company", "contract_hours", "contract_type"]]

    employees["id"] = employees["id"].astype(int)
    employees["contract_hours"] = employees["contract_hours"].astype(float)
    employees = employees[employees["contract_hours"] > 0]
    employees = employees.drop_duplicates(subset=["id"]).reset_index(drop=True)

    return employees


# ---------------------------------------------------------------------------
# Combined employee + rate template
# ---------------------------------------------------------------------------
# Rather than parsing the raw medewerkers file and a separate uurprijs file
# at every run, the planner now uses ONE editable template that Rosa
# maintains. It is generated once from the medewerkers file (pre-filled),
# then Rosa fills in the hourly_cost column and, optionally, an end_week for
# anyone leaving mid-horizon. This removes the dependency on the fragile
# uurprijs format and keeps all employee data in one place.

EMPLOYEE_TEMPLATE_COLUMNS = {
    # required
    "id":             "Employee number (from the staff system)",
    "role":           "Job role (Combicoach, Stagiair, Manager, ...)",
    "company":        "Combibrug or CC",
    "contract_hours": "Contracted hours per week",
    "hourly_cost":    "Cost to Combibrug per hour, € (including employer overhead)",
    # optional
    "contract_type":  "permanent / fixed_term / stage / oproep (informational)",
    "end_week":       "Last week the employee is available (blank = whole horizon)",
    "is_dreammaker":  "TRUE if this employee is a Dreammaker (Combiworld/MDT), else blank",
}


def generate_employee_template(medewerkers_path, out_path,
                               planning_date="2026-01-01",
                               default_rate=22.0):
    """Create a pre-filled employee template from the medewerkers file.

    Rosa runs this once, then edits the resulting Excel file: she fills in
    the real hourly_cost per person, marks any Dreammakers, and sets an
    end_week for anyone leaving mid-horizon. After that the planner only
    ever reads the edited template — the raw medewerkers file is no longer
    needed.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    emp = load_employees(medewerkers_path, planning_date=planning_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = list(EMPLOYEE_TEMPLATE_COLUMNS.keys())
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    inst_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=ci, value=EMPLOYEE_TEMPLATE_COLUMNS[h])
        c.fill = inst_fill
        c.font = Font(italic=True, color="1F4E79")

    # Pre-fill rows from the medewerkers data
    for ri, (_, e) in enumerate(emp.iterrows(), start=3):
        ws.cell(row=ri, column=1, value=int(e["id"]))
        ws.cell(row=ri, column=2, value=e["role"])
        ws.cell(row=ri, column=3, value=e["company"])
        ws.cell(row=ri, column=4, value=float(e["contract_hours"]))
        ws.cell(row=ri, column=5, value=default_rate)        # hourly_cost — Rosa edits
        ws.cell(row=ri, column=6, value=e["contract_type"])
        ws.cell(row=ri, column=7, value="")                  # end_week — blank
        ws.cell(row=ri, column=8, value="")                  # is_dreammaker — blank

    widths = [10, 18, 12, 16, 30, 16, 12, 16]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    wb.save(out_path)


def load_employee_template(path_or_buffer):
    """Load the edited employee template (one combined file with rates)."""
    try:
        df = pd.read_excel(path_or_buffer)
    except Exception as e:
        raise DataLoadError(f"Could not open the employee template: {e}")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    # Drop the instruction row if present (its 'id' cell isn't a number)
    df = df[pd.to_numeric(df["id"], errors="coerce").notna()].copy()

    required = ["id", "role", "company", "contract_hours", "hourly_cost"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLoadError(
            f"The employee template is missing required columns: {missing}. "
            f"Found: {list(df.columns)}"
        )

    df["id"] = pd.to_numeric(df["id"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["id"])
    df["id"] = df["id"].astype(int)
    df["contract_hours"] = pd.to_numeric(df["contract_hours"], errors="coerce").fillna(0).astype(float)
    df["hourly_cost"] = pd.to_numeric(df["hourly_cost"], errors="coerce")

    # company normalisation
    df["company"] = df["company"].apply(
        lambda v: "CC" if ("cc" in str(v).lower() or "combicoach" in str(v).lower()) else "Combibrug"
    )

    if "contract_type" not in df.columns:
        df["contract_type"] = "unknown"

    # end_week: optional last available week
    if "end_week" not in df.columns:
        df["end_week"] = pd.NA
    else:
        df["end_week"] = pd.to_numeric(df["end_week"], errors="coerce").astype("Int64")

    # is_dreammaker: optional boolean
    if "is_dreammaker" not in df.columns:
        df["is_dreammaker"] = False
    else:
        df["is_dreammaker"] = df["is_dreammaker"].apply(
            lambda v: str(v).strip().lower() in ("true", "1", "yes", "ja", "x")
        )

    df = df[df["contract_hours"] > 0].drop_duplicates(subset=["id"]).reset_index(drop=True)

    # Any missing hourly_cost falls back to a flag the dashboard can surface
    missing_rate = df["hourly_cost"].isna()
    df["cost_source"] = "template"
    df.loc[missing_rate, "cost_source"] = "missing"

    out_cols = ["id", "role", "company", "contract_hours", "hourly_cost",
                "contract_type", "end_week", "is_dreammaker", "cost_source"]
    return df[out_cols]
    """Load the locations file and return a cleaned locations DataFrame.

    Steps:
      1. Read the wide-format sheet (employees x locations, HH:MM cells).
      2. Melt into long format → one row per (employee, location) with hours.
      3. Aggregate to weekly demand per location.
      4. Extract duration metadata: text in row 4 first, cell colours as fallback.
    """
    # --- 2a. Read and melt
    try:
        sheet_names = pd.ExcelFile(locations_path).sheet_names
    except FileNotFoundError:
        raise DataLoadError(f"Locations file not found: {locations_path}")
    except Exception as e:
        raise DataLoadError(f"Could not open the locations file: {e}")

    if "Worksheet" not in sheet_names:
        raise DataLoadError(
            f"The locations file does not contain a sheet called 'Worksheet'. "
            f"Sheets found: {sheet_names}. "
            f"Rename the data sheet to 'Worksheet' or adjust load_locations()."
        )

    raw = pd.read_excel(locations_path, sheet_name="Worksheet", header=2)
    first_col = raw.columns[0]
    raw = raw.rename(columns={first_col: "employee_id"})
    raw = raw.dropna(subset=["employee_id"])
    raw = raw[pd.to_numeric(raw["employee_id"], errors="coerce").notna()]
    raw["employee_id"] = raw["employee_id"].astype(int)

    summary_cols = [
        "Gewerkte uren", "Gewerkte dagen", "Verlof opgenomen",
        "Ziekte-uren", "Bijzonder verlof", "Totaal maaltijden",
    ]
    loc_cols = [
        c for c in raw.columns
        if c not in ["employee_id"] + summary_cols
        and not str(c).startswith("Overige teams")
    ]

    long_df = raw[["employee_id"] + loc_cols].melt(
        id_vars=["employee_id"],
        var_name="location_full",
        value_name="hours_str",
    )
    long_df["hours"] = long_df["hours_str"].apply(hhmm_to_hours)
    long_df = long_df[long_df["hours"] > 0]

    # Split "Combibrug CC - De Dukdalf" → prefix "Combibrug CC", location "De Dukdalf"
    # The prefix is used only to derive the company; it is NOT kept in the
    # final DataFrame.
    def split_loc(name):
        s = str(name)
        if " - " in s:
            prefix, loc = s.split(" - ", 1)
            return prefix.strip(), loc.strip()
        return "Unknown", s.strip()

    long_df[["_company_prefix", "_site_part"]] = long_df["location_full"].apply(
        lambda x: pd.Series(split_loc(x))
    )
    long_df["company"] = long_df["_company_prefix"].apply(
        lambda d: "CC" if "CC" in d.upper() else "Combibrug"
    )

    # --- 2b. Extract duration metadata from the workbook directly.
    # Priority: explicit text in row 4 > cell colour in row 3 > default 12 months.
    # Text is preferred because it survives copy-pasting, CSV export, and
    # editing in other tools (Google Sheets, LibreOffice). Colours are kept
    # as a fallback so the original coloured file still works without changes.
    wb = load_workbook(locations_path, data_only=True)
    ws = wb["Worksheet"]
    meta = []
    for col_idx in range(8, ws.max_column + 1):
        name_cell = ws.cell(row=3, column=col_idx)
        duration_cell = ws.cell(row=4, column=col_idx)
        if not name_cell.value:
            continue
        name = str(name_cell.value).strip()
        if name.startswith("Overige teams"):
            continue

        # 1. Try text in row 4 first.
        dur, source = _parse_duration_text(duration_cell.value)

        # 2. Fall back to colour fill on row 3.
        if dur is None:
            colour = None
            if name_cell.fill and name_cell.fill.start_color:
                rgb = name_cell.fill.start_color.rgb
                if isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                    colour = rgb
            if colour in COLOR_TO_DURATION:
                dur = COLOR_TO_DURATION[colour]
                source = f"color:{colour}"

        # 3. Default.
        if dur is None:
            dur = 12
            source = "default"

        meta.append({
            "location_full": name,
            "duration_months": dur,
            "duration_source": source,
        })
    meta_df = pd.DataFrame(meta)

    # --- 2c. Aggregate to weekly demand and join with metadata
    demand = (long_df
              .groupby(["location_full", "company"])["hours"]
              .sum()
              .reset_index())
    demand["weekly_hours"] = (demand["hours"] / WEEKS_IN_FILE).round(1)

    locations = demand.merge(meta_df, on="location_full", how="left")
    locations = locations.rename(columns={"location_full": "site"})
    locations["duration_months"] = locations["duration_months"].fillna(12).astype(int)
    locations["duration_source"] = locations["duration_source"].fillna("default")

    # Historical files don't carry a separate project name — the column
    # header in the source spreadsheet IS the site name. Treat each site as
    # its own one-site project.
    locations["project"] = locations["site"]

    # Convert to weeks for the multi-period MILP, and assume every project
    # starts at week 1 of the horizon (the historical file has no start info).
    locations["duration_weeks"] = (locations["duration_months"] * 52.0 / 12.0).round().astype(int)
    locations["start_week"] = 1

    # Historical files don't have Four-Eyes / project-leader info — leave the
    # corresponding columns at their inert defaults so the model treats them
    # as "no requirement".
    locations["min_headcount"] = 0
    locations["project_leader_id"] = pd.NA

    # Default: no per-role headcount requirement (the MILP falls back to total
    # weekly hours as the binding demand constraint until Rosa provides headcount).
    for r in ALL_ROLES:
        locations[f"req_{r}"] = 0

    locations = locations[locations["weekly_hours"] > 0].reset_index(drop=True)
    return locations


# ---------------------------------------------------------------------------
# Step 3 — load the per-employee hourly rate file
# ---------------------------------------------------------------------------

def load_rates(rates_path, sheet_name="Blad1"):
    """Load the per-employee hourly rate file (Berekening uurloon).

    The file has one row per employee and 12 monthly blocks across,
    each block being four columns:

        [base €/h]  [€/h incl. werkgeverslasten]  [hours that month]  [monthly cost]

    Rates can change month to month (CAO increases, raises, etc.). Empty
    cells indicate the employee was not active that month. Some cells
    contain the literal text "stage" — these are stagiaires whose hourly
    cost is treated separately.

    Returns
    -------
    DataFrame with columns:
        id           int      employee id (matches medewerkers.Nr)
        rate_base    float    average base €/h across non-empty months
        rate_incl    float    average €/h including employer costs
        n_months     int      how many months contributed to the average
        is_stage     bool     True if every non-empty cell said "stage"
    """
    try:
        xl = pd.ExcelFile(rates_path)
    except FileNotFoundError:
        raise DataLoadError(f"Rates file not found: {rates_path}")
    except Exception as e:
        raise DataLoadError(f"Could not open the rates file: {e}")

    if sheet_name not in xl.sheet_names:
        raise DataLoadError(
            f"The rates file does not contain a sheet called '{sheet_name}'. "
            f"Sheets found: {xl.sheet_names}."
        )

    # The real header is on row 2 (Excel 1-indexed) → pandas header=1
    raw = pd.read_excel(rates_path, sheet_name=sheet_name, header=1)

    if "Voornaam" not in raw.columns:
        raise DataLoadError(
            "The rates file is missing the 'Voornaam' column (which holds "
            "the anonymised employee id)."
        )

    # Identify the 12 base-rate columns: the ones whose header is a 2025 date.
    # pandas may return either datetime.datetime or pd.Timestamp depending on
    # the openpyxl/pandas version — accept both.
    import datetime as _dt
    base_rate_cols = []
    for i, c in enumerate(raw.columns):
        if isinstance(c, (pd.Timestamp, _dt.datetime)) and c.year == 2025:
            base_rate_cols.append(i)
    if len(base_rate_cols) == 0:
        raise DataLoadError(
            "Could not find any monthly rate columns (expected dated headers "
            "like 2025-01-01 in row 2)."
        )

    # The 'incl werkgeverslasten' column for each month is immediately to the
    # right of the base-rate column.
    incl_rate_cols = [c + 1 for c in base_rate_cols]

    rows = []
    for _, r in raw.iterrows():
        emp_id = r["Voornaam"]
        if pd.isna(emp_id):
            continue
        try:
            emp_id = int(emp_id)
        except (TypeError, ValueError):
            continue  # skip non-numeric ids

        base_vals, incl_vals, n_non_empty, stage_flags = [], [], 0, []
        for bi, ii in zip(base_rate_cols, incl_rate_cols):
            v_base = r.iloc[bi]
            v_incl = r.iloc[ii] if ii < len(r) else None
            if pd.isna(v_base) or v_base == "" or v_base == 0:
                continue
            if isinstance(v_base, str):
                if v_base.strip().lower() == "stage":
                    stage_flags.append(True)
                    n_non_empty += 1
                    continue
                else:
                    # Unrecognised text — skip
                    continue
            base_vals.append(float(v_base))
            try:
                incl_vals.append(float(v_incl) if not pd.isna(v_incl) else float(v_base))
            except (TypeError, ValueError):
                incl_vals.append(float(v_base))
            stage_flags.append(False)
            n_non_empty += 1

        if n_non_empty == 0:
            continue

        is_stage = len(stage_flags) > 0 and all(stage_flags)
        rows.append({
            "id": emp_id,
            "rate_base": round(sum(base_vals) / len(base_vals), 2) if base_vals else None,
            "rate_incl": round(sum(incl_vals) / len(incl_vals), 2) if incl_vals else None,
            "n_months": n_non_empty,
            "is_stage": is_stage,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # Employees occasionally appear twice in the rates file (mid-year role
    # changes, department transfers). Keep the row with the most months of
    # data — it's the more representative one. Stagiair rows are preferred
    # only if no non-stage row exists for the same id.
    df = (df
          .sort_values(["id", "is_stage", "n_months"],
                       ascending=[True, True, False])  # non-stage first, then most months
          .drop_duplicates(subset=["id"], keep="first")
          .reset_index(drop=True))
    return df


# ---------------------------------------------------------------------------
# Step 4 — load a forward-looking demand template (simple format)
# ---------------------------------------------------------------------------

DEMAND_TEMPLATE_COLUMNS = {
    # required
    "site":           "Name of the specific site/location where work happens",
    "company":        "Which company runs it: 'Combibrug' or 'CC'",
    "weekly_hours":   "Expected hours of work needed per week (decimal)",
    # optional — sensible defaults applied when missing
    "project":            "Project the site belongs to (default = same as site name)",
    "duration_months":    "How many months the project runs (default = full horizon)",
    "start_week":         "Which week of the horizon the project starts in, 1-indexed (default 1)",
    "min_headcount":      "Minimum number of distinct people required per week (default 0, i.e. no requirement). Use 2 for Combiworld/MDT (Four-Eyes rule).",
    "project_leader_id":  "Employee ID of a designated project leader who must work at this site every active week (default: no fixed leader).",
    "notes":              "Free-text notes — not used by the model",
}

def load_demand_template(path_or_buffer):
    """Load a forward-looking demand template file.

    Accepts either a file path (str / Path) or a file-like object
    (e.g. a Streamlit UploadedFile). Supports both .xlsx and .csv.

    The template has one row per project location and only needs a
    handful of columns — see DEMAND_TEMPLATE_COLUMNS for the full list.
    This replaces the historical wide-format locations file for future
    planning cycles, where actual timesheets don't exist yet.

    Returns a locations DataFrame in the same shape as load_locations(),
    so the rest of the app works identically regardless of which format
    was used.
    """
    # --- Detect format and read
    try:
        # Check if it's a file-like object with a name attribute (UploadedFile)
        name = getattr(path_or_buffer, "name", str(path_or_buffer))
        if str(name).lower().endswith(".csv"):
            df = pd.read_csv(path_or_buffer)
        else:
            df = pd.read_excel(path_or_buffer)
    except Exception as e:
        raise DataLoadError(f"Could not read the demand template: {e}")

    # --- Normalise column names: strip spaces, lowercase
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    # --- Backward compatibility: older templates used 'location' as the
    # column name. New templates use 'site' (a more accurate term — a site
    # is the physical place where work happens, while a project is the
    # funded activity, which can span multiple sites). Accept either.
    if "site" not in df.columns and "location" in df.columns:
        df = df.rename(columns={"location": "site"})

    # --- Check required columns
    required = ["site", "company", "weekly_hours"]
    missing_cols = [c for c in required if c not in df.columns]
    if missing_cols:
        raise DataLoadError(
            f"The demand template is missing required columns: {missing_cols}. "
            f"Columns found: {list(df.columns)}. "
            f"Required: {required} (the old name 'location' is also accepted for 'site')."
        )

    # --- Validate and clean
    df = df.dropna(subset=["site", "weekly_hours"])
    df["site"] = df["site"].astype(str).str.strip()
    df["weekly_hours"] = pd.to_numeric(df["weekly_hours"], errors="coerce").fillna(0)
    df = df[df["weekly_hours"] > 0].copy()

    if df.empty:
        raise DataLoadError(
            "The demand template has no valid rows. "
            "Check that 'weekly_hours' contains numbers greater than 0."
        )

    # --- Apply defaults for optional columns
    if "duration_months" not in df.columns:
        df["duration_months"] = 12
    else:
        df["duration_months"] = pd.to_numeric(
            df["duration_months"], errors="coerce"
        ).fillna(12).astype(int)

    # duration_weeks: convert duration_months using the standard 52/12 ratio.
    # This is the unit the multi-period MILP uses.
    df["duration_weeks"] = (df["duration_months"] * 52.0 / 12.0).round().astype(int)

    # start_week: 1-indexed (week 1 = first week of the planning horizon).
    # Default 1 means "starts at the beginning of the horizon".
    if "start_week" not in df.columns:
        df["start_week"] = 1
    else:
        df["start_week"] = (pd.to_numeric(df["start_week"], errors="coerce")
                            .fillna(1).astype(int).clip(lower=1))

    # project: optional column grouping sites into named projects. If a row
    # has no project specified, the site is its own one-site project.
    if "project" not in df.columns:
        df["project"] = df["site"]
    else:
        # Cells that are blank, NaN, or whitespace fall back to the site name.
        df["project"] = df["project"].astype(str).str.strip()
        df.loc[df["project"].isin(["", "nan", "None"]), "project"] = (
            df.loc[df["project"].isin(["", "nan", "None"]), "site"]
        )

    # min_headcount: minimum number of distinct people required per active
    # week (the Four-Eyes rule for Combiworld/MDT, etc). Default 0 means
    # no requirement. Coerced to int >= 0.
    if "min_headcount" not in df.columns:
        df["min_headcount"] = 0
    else:
        df["min_headcount"] = (pd.to_numeric(df["min_headcount"], errors="coerce")
                               .fillna(0).astype(int).clip(lower=0))

    # project_leader_id: optional employee ID. If set, that specific
    # employee is forced to work at the site every active week.
    # Stored as a nullable int — pandas Int64 supports NaN cleanly.
    if "project_leader_id" not in df.columns:
        df["project_leader_id"] = pd.NA
    else:
        df["project_leader_id"] = pd.to_numeric(
            df["project_leader_id"], errors="coerce"
        ).astype("Int64")

    # Normalise company: accept 'cc', 'combicoach', 'combibrug', 'cb' etc.
    def _normalise_company(v):
        s = str(v).strip().lower()
        if "cc" in s or "combicoach" in s:
            return "CC"
        return "Combibrug"

    df["company"] = df["company"].apply(_normalise_company)

    # Add the per-role requirement columns (all zero — no headcount data yet)
    for r in ALL_ROLES:
        df[f"req_{r}"] = 0

    # Mark where the duration came from so the Inputs tab can show it
    df["duration_source"] = "template"

    # Keep only the columns the rest of the app expects
    out_cols = [
        "site", "project", "company",
        "weekly_hours", "duration_months", "duration_weeks", "start_week",
        "min_headcount", "project_leader_id",
        "duration_source",
    ] + [f"req_{r}" for r in ALL_ROLES]

    return df[out_cols].reset_index(drop=True)


def generate_demand_template_example(path):
    """Write a blank-but-filled example demand template to path.

    Rosa can use this as a starting point for each new planning cycle.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = [
        # site, project, company, weekly_hours, duration_months, start_week, min_headcount, project_leader_id, notes
        # CC sites — no Four-Eyes requirement, no fixed leader needed
        ("Olympiaplein",                 "Olympiaplein",         "CC",         18, 12, 1, 0, None, "CC, year-round sports location"),
        ("Stadionplein",                 "Stadionplein",         "CC",         14, 12, 1, 0, None, "CC, year-round sports location"),
        ("Apollolaan",                   "Apollolaan",           "CC",         10, 12, 1, 0, None, "CC, year-round sports location"),
        ("Mercatorplein",                "Mercatorplein",        "CC",         16, 12, 1, 0, None, "CC, year-round sports location"),
        ("Surinameplein",                "Surinameplein",        "CC",         12, 12, 1, 0, None, "CC, year-round sports location"),
        # Combiworld project — runs at three different sites. Four-Eyes rule:
        # each Combiworld site needs at least 2 people per active week.
        ("Park Frankendael",             "Combiworld-Oost",      "Combibrug",  12, 10, 1, 2, None, "Combiworld — Four-Eyes (min 2 people)"),
        ("Sportpark Drieburg",           "Combiworld-Oost",      "Combibrug",   8, 10, 1, 2, None, "Combiworld — Four-Eyes (min 2 people)"),
        ("Bouwkeet Indische Buurt",      "Combiworld-Oost",      "Combibrug",   6, 10, 1, 2, None, "Combiworld — Four-Eyes (min 2 people)"),
        # MDT projects — also Four-Eyes
        ("Buurtcentrum Watergraafsmeer", "MDT-Watergraafsmeer",  "Combibrug",   6,  4, 1, 2, None, "MDT subsidy, 4 months, Four-Eyes"),
        ("Buurtcentrum Diemen",          "MDT-Diemen",           "Combibrug",   5,  4, 5, 2, None, "MDT subsidy, starts a month later"),
        # BSC example with a fixed project leader (employee ID would be filled in by Rosa)
        ("BSC Oost",                     "BSC-Oost",             "Combibrug",   8, 10, 1, 0, None, "BSC — fill in project_leader_id with the lead's employee ID"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demand template"

    # Header row
    headers = ["site", "project", "company", "weekly_hours",
               "duration_months", "start_week",
               "min_headcount", "project_leader_id", "notes"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Instruction row (light blue background)
    instructions = [
        "Physical site / location name",
        "Project name (leave blank if same as site)",
        "'Combibrug' or 'CC'",
        "Hours/week (decimal)",
        "Months project runs",
        "Start week (1 = first week)",
        "Min people/week (2 for Combiworld/MDT, 0 otherwise)",
        "Employee ID of fixed project leader (or blank)",
        "Any notes (not used by model)",
    ]
    inst_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                            fill_type="solid")
    for col_idx, inst in enumerate(instructions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=inst)
        cell.fill = inst_fill
        cell.font = Font(italic=True, color="1F4E79")

    # Data rows — replace Python None with empty string so Excel cells stay blank
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx,
                    value="" if val is None else val)

    # Column widths
    widths = [28, 22, 12, 14, 17, 14, 16, 19, 45]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = w

    wb.save(path)


# ---------------------------------------------------------------------------
# Convenience: load all files at once
# ---------------------------------------------------------------------------

def load_all(employees_path, locations_path, rates_path=None,
             planning_date="2026-01-01"):
    """Load employees, locations, and (optionally) rates.

    locations_path can be either:
      - The historical wide-format Excel file (Kopie van locaties...)
      - A forward-looking demand template (.xlsx or .csv)

    The function auto-detects which format it is by checking whether the
    file contains a 'weekly_hours' column. If it does, it's treated as a
    demand template; otherwise the full historical parser is used.

    Returns (employees_df, locations_df, rates_df).
    """
    employees = load_employees(employees_path, planning_date=planning_date)

    # Auto-detect locations format
    try:
        probe = pd.read_excel(locations_path, nrows=1)
        probe.columns = [str(c).strip().lower().replace(" ", "_")
                         for c in probe.columns]
        is_template = "weekly_hours" in probe.columns
    except Exception:
        is_template = False

    if is_template:
        locations = load_demand_template(locations_path)
    else:
        locations = load_locations(locations_path)

    rates = load_rates(rates_path) if rates_path else None
    return employees, locations, rates
