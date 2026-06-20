"""
data_loader.py
--------------
Reads the two Combibrug planning templates and returns cleaned DataFrames:
    - employees_df: one row per employee (internal staff and freelancers)
    - locations_df: one row per project site with weekly demand

The MILP and the UI only ever consume these two templates. There is no
dependency on any raw company export — both templates are filled in
directly by the planner.
"""

import pandas as pd


ALL_ROLES = [
    "Combicoach", "Stagiair", "Jongerenbegeleider", "Buurtsportcoach",
    "Manager", "Cultuurcoordinator", "Leefstijlcoach", "Other",
]


class DataLoadError(Exception):
    """Raised when an input file cannot be parsed.

    Use a clear, planner-friendly message — these are surfaced directly
    in the Streamlit UI.
    """
    pass


def _check_required_columns(df, required, file_label):
    """Raise DataLoadError if any required columns are missing."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLoadError(
            f"The {file_label} file is missing required columns: {missing}. "
            f"Columns found: {list(df.columns)[:10]}{'...' if len(df.columns) > 10 else ''}"
        )


DEMAND_TEMPLATE_COLUMNS = {
    "site":           "Name of the specific site/location where work happens",
    "company":        "Which company runs it: 'Combibrug' or 'CC'",
    "weekly_hours":   "Expected hours of work needed per week (decimal)",
    "project":            "Project the site belongs to (default = same as site name)",
    "duration_months":    "How many months the project runs (default = full horizon)",
    "start_week":         "Which week of the horizon the project starts in, 1-indexed (default 1)",
    "min_headcount":      "Minimum number of distinct people required per week (default 0, i.e. no requirement). Use 2 for Combiworld/MDT (Four-Eyes rule).",
    "project_leader_id":  "Employee ID of a designated project leader who must work at this site every active week (default: no fixed leader).",
    "available_mon":      "Time windows the site is available on Monday, e.g. '08:00-16:00' or '08:00-12:00; 13:00-17:00'. Blank = closed.",
    "available_tue":      "Time windows the site is available on Tuesday.",
    "available_wed":      "Time windows the site is available on Wednesday.",
    "available_thu":      "Time windows the site is available on Thursday.",
    "available_fri":      "Time windows the site is available on Friday.",
    "notes":              "Free-text notes — not used by the model",
}

DAY_COLS = ["available_mon", "available_tue", "available_wed",
            "available_thu", "available_fri"]
DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri"]


def _parse_windows(cell):
    """Parse a day-availability cell into a list of (start_h, end_h) tuples
    in decimal hours.

    Examples:
        "08:00-16:00"               -> [(8.0, 16.0)]
        "08:00-12:00; 13:00-17:00"  -> [(8.0, 12.0), (13.0, 17.0)]
        "" or NaN                   -> []

    The parser tolerates spaces, en-dashes vs hyphens, and both ':' and '.'
    as the H/M separator. Invalid pieces are skipped silently.
    """
    if cell is None:
        return []
    s = str(cell).strip()
    if s == "" or s.lower() in ("nan", "none"):
        return []

    s = s.replace("–", "-").replace("—", "-")
    out = []
    for piece in s.split(";"):
        piece = piece.strip()
        if not piece:
            continue
        if "-" not in piece:
            continue
        try:
            a, b = piece.split("-", 1)
            def _to_h(t):
                t = t.strip().replace(".", ":")
                if ":" not in t:
                    return float(t)
                hh, mm = t.split(":", 1)
                return int(hh) + int(mm) / 60.0
            start_h = _to_h(a)
            end_h   = _to_h(b)
            if end_h > start_h and 0 <= start_h < 24 and 0 < end_h <= 24:
                out.append((start_h, end_h))
        except (ValueError, IndexError):
            continue
    return out

def load_demand_template(path_or_buffer):
    """Load a forward-looking demand template file.

    Accepts either a file path (str / Path) or a file-like object
    (e.g. a Streamlit UploadedFile). Supports both .xlsx and .csv.

    The template has one row per project location and only needs a
    handful of columns — see DEMAND_TEMPLATE_COLUMNS for the full list.
    This replaces the historical wide-format locations file for future
    planning cycles, where actual timesheets don't exist yet.

    Returns a cleaned locations DataFrame ready for the optimisation model.
    so the rest of the app works identically regardless of which format
    was used.
    """
    try:
        name = getattr(path_or_buffer, "name", str(path_or_buffer))
        if str(name).lower().endswith(".csv"):
            df = pd.read_csv(path_or_buffer)
        else:
            df = pd.read_excel(path_or_buffer)
    except Exception as e:
        raise DataLoadError(f"Could not read the demand template: {e}")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    if "site" not in df.columns and "location" in df.columns:
        df = df.rename(columns={"location": "site"})

    required = ["site", "company", "weekly_hours"]
    missing_cols = [c for c in required if c not in df.columns]
    if missing_cols:
        raise DataLoadError(
            f"The demand template is missing required columns: {missing_cols}. "
            f"Columns found: {list(df.columns)}. "
            f"Required: {required} (the old name 'location' is also accepted for 'site')."
        )

    df = df.dropna(subset=["site", "weekly_hours"])
    df["site"] = df["site"].astype(str).str.strip()
    df["weekly_hours"] = pd.to_numeric(df["weekly_hours"], errors="coerce").fillna(0)
    df = df[df["weekly_hours"] > 0].copy()

    if df.empty:
        raise DataLoadError(
            "The demand template has no valid rows. "
            "Check that 'weekly_hours' contains numbers greater than 0."
        )

    if "duration_months" not in df.columns:
        df["duration_months"] = 12
    else:
        df["duration_months"] = pd.to_numeric(
            df["duration_months"], errors="coerce"
        ).fillna(12).astype(int)

    df["duration_weeks"] = (df["duration_months"] * 52.0 / 12.0).round().astype(int)

    if "start_week" not in df.columns:
        df["start_week"] = 1
    else:
        df["start_week"] = (pd.to_numeric(df["start_week"], errors="coerce")
                            .fillna(1).astype(int).clip(lower=1))

    if "project" not in df.columns:
        df["project"] = df["site"]
    else:
        df["project"] = df["project"].astype(str).str.strip()
        df.loc[df["project"].isin(["", "nan", "None"]), "project"] = (
            df.loc[df["project"].isin(["", "nan", "None"]), "site"]
        )

    if "min_headcount" not in df.columns:
        df["min_headcount"] = 0
    else:
        df["min_headcount"] = (pd.to_numeric(df["min_headcount"], errors="coerce")
                               .fillna(0).astype(int).clip(lower=0))

    if "project_leader_id" not in df.columns:
        df["project_leader_id"] = pd.NA
    else:
        df["project_leader_id"] = pd.to_numeric(
            df["project_leader_id"], errors="coerce"
        ).astype("Int64")

    def _normalise_company(v):
        s = str(v).strip().lower()
        if "cc" in s or "combicoach" in s:
            return "CC"
        return "Combibrug"

    df["company"] = df["company"].apply(_normalise_company)

    for r in ALL_ROLES:
        df[f"req_{r}"] = 0

    df["duration_source"] = "template"

    parsed_windows = []
    totals = []
    for _, row in df.iterrows():
        site_windows = []
        for col, day in zip(DAY_COLS, DAY_NAMES):
            if col in df.columns:
                for (s_h, e_h) in _parse_windows(row.get(col)):
                    site_windows.append((day, s_h, e_h))
        parsed_windows.append(site_windows)
        totals.append(round(sum(e - s for (_, s, e) in site_windows), 2))
    df["windows"] = parsed_windows
    df["total_window_hours"] = totals

    out_cols = [
        "site", "project", "company",
        "weekly_hours", "duration_months", "duration_weeks", "start_week",
        "min_headcount", "project_leader_id",
        "duration_source",
        "windows", "total_window_hours",
    ]
    for c in DAY_COLS:
        if c in df.columns:
            out_cols.append(c)
    out_cols += [f"req_{r}" for r in ALL_ROLES]

    return df[out_cols].reset_index(drop=True)


def generate_demand_template_example(path):
    """Write a blank-but-filled example demand template to path.

    Rosa can use this as a starting point for each new planning cycle.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = [
        ("Dukdalf",                "Dukdalf",              "CC",    8, 12, 1, 0, None,
         "", "08:00-16:00", "", "", "",
         "CC site, Tuesdays only"),
        ("Ley",                    "Ley",                  "CC",   16, 12, 1, 0, None,
         "08:00-13:30", "08:00-13:30", "08:00-13:30", "", "",
         "CC site, Mon-Wed mornings"),
        ("Apollo",                 "Apollo",               "CC",    8, 12, 1, 0, None,
         "", "", "08:00-16:30", "", "",
         "CC site, Wednesdays only"),
        ("Olympiaplein",           "Olympiaplein",         "CC",   16, 12, 1, 0, None,
         "09:00-17:00", "09:00-17:00", "", "", "",
         "CC year-round sports location"),
        ("Mercatorplein",          "Mercatorplein",        "CC",   16, 12, 1, 0, None,
         "", "09:00-17:00", "09:00-17:00", "", "",
         "CC year-round sports location"),
        ("Park Frankendael",       "Combiworld-Oost",      "Combibrug", 12, 10, 1, 2, None,
         "09:00-15:00", "09:00-15:00", "09:00-15:00", "", "",
         "Combiworld — Four-Eyes (min 2 people)"),
        ("Sportpark Drieburg",     "Combiworld-Oost",      "Combibrug",  8, 10, 1, 2, None,
         "", "13:00-17:00", "13:00-17:00", "", "",
         "Combiworld — Four-Eyes"),
        ("Bouwkeet Indische Buurt","Combiworld-Oost",      "Combibrug",  6, 10, 1, 2, None,
         "", "", "", "13:00-17:00", "13:00-17:00",
         "Combiworld — Four-Eyes"),
        ("Buurtcentrum Watergraafsmeer","MDT-Watergraafsmeer","Combibrug",  6,  4, 1, 2, None,
         "09:00-13:00", "13:00-17:00", "", "", "",
         "MDT subsidy, 4 months, Four-Eyes"),
        ("Buurtcentrum Diemen",    "MDT-Diemen",           "Combibrug",   5,  4, 5, 2, None,
         "09:00-13:00", "", "13:00-17:00", "", "",
         "MDT subsidy, starts later"),
        ("BSC Oost",               "BSC-Oost",             "Combibrug",   8, 10, 1, 0, None,
         "09:00-13:00", "", "09:00-13:00", "", "",
         "BSC — fill in project_leader_id with the lead's employee ID"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demand template"

    headers = ["site", "project", "company", "weekly_hours",
               "duration_months", "start_week",
               "min_headcount", "project_leader_id",
               "available_mon", "available_tue", "available_wed",
               "available_thu", "available_fri",
               "notes"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    instructions = [
        "Physical site / location name",
        "Project name (leave blank if same as site)",
        "'Combibrug' or 'CC'",
        "Hours/week (decimal)",
        "Months project runs",
        "Start week (1 = first week)",
        "Min people/week (2 for Combiworld/MDT, 0 otherwise)",
        "Employee ID of fixed project leader (or blank)",
        "Mon time windows, e.g. '08:00-16:00' or '08:00-12:00; 13:00-17:00'",
        "Tue time windows (blank = closed)",
        "Wed time windows",
        "Thu time windows",
        "Fri time windows",
        "Any notes (not used by model)",
    ]
    inst_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                            fill_type="solid")
    for col_idx, inst in enumerate(instructions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=inst)
        cell.fill = inst_fill
        cell.font = Font(italic=True, color="1F4E79")

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx,
                    value="" if val is None else val)

    widths = [28, 22, 12, 14, 17, 14, 16, 19, 18, 18, 18, 18, 18, 45]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = w

    wb.save(path)


EMPLOYEE_TEMPLATE_COLUMNS = {
    "id":             "Employee number (from the staff system)",
    "role":           "Job role (Combicoach, Stagiair, Manager, ...). For freelancers, use the role they would cover.",
    "company":        "Combibrug or CC",
    "contract_hours": "Contracted hours per week (for freelancers: a practical weekly ceiling, e.g. 40)",
    "hourly_cost":    "Cost to Combibrug per hour, € — REQUIRED for every row, including freelancers",
    "contract_type":  "permanent / fixed_term / stage / oproep / freelancer",
    "end_week":       "Last week the employee is available (blank = whole horizon)",
    "is_dreammaker":  "TRUE if this employee is a Dreammaker (Combiworld/MDT), else blank",
}

# Contract types that count as "freelancer" for eligibility and cost-priority
# purposes. Freelancers participate in the MILP exactly like internal staff
# (same eligibility, same daily-scheduling rules) but are excluded from the
# minimum-utilisation target (C12) and are the last resort once internal
# staff and other freelancers cannot cover a site's demand — this ordering
# happens naturally through the objective as long as freelancer hourly_cost
# is set higher than internal staff, which the planner controls via the
# template, not the code.
FREELANCER_CONTRACT_TYPES = {"freelancer", "zzp"}


def generate_employee_template(out_path):
    """Write a blank-but-illustrative employee template to out_path.

    The template is filled in directly by the planner — there is no raw
    company export to seed it from. A few example rows (including one
    freelancer) are included to show the expected format; the planner
    deletes or overwrites them.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    example_rows = [
        (1,  "Combicoach",     "CC",        32, 28.0, "permanent",  "", ""),
        (2,  "Buurtsportcoach","Combibrug", 36, 30.0, "permanent",  "", ""),
        (3,  "Stagiair",       "Combibrug", 16, 14.0, "stage",      "", ""),
        (4,  "Combicoach",     "CC",        24, 22.0, "fixed_term", 8, ""),
        (5,  "Buurtsportcoach","Combibrug", 32, 45.0, "freelancer", "", ""),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = list(EMPLOYEE_TEMPLATE_COLUMNS.keys())
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    inst_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                            fill_type="solid")
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=ci, value=EMPLOYEE_TEMPLATE_COLUMNS[h])
        c.fill = inst_fill
        c.font = Font(italic=True, color="1F4E79")

    for ri, row in enumerate(example_rows, start=3):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value="" if val is None else val)

    widths = [10, 18, 12, 16, 30, 16, 12, 16]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    wb.save(out_path)


def load_employee_template(path_or_buffer):
    """Load the employee template (internal staff and freelancers, one file).

    hourly_cost is required for every row — there is no fallback rate.
    Rows with a missing hourly_cost raise a DataLoadError listing the
    offending employee IDs so the planner can fix the template directly.
    """
    try:
        df = pd.read_excel(path_or_buffer)
    except Exception as e:
        raise DataLoadError(f"Could not open the employee template: {e}")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    df = df[pd.to_numeric(df["id"], errors="coerce").notna()].copy()

    required = ["id", "role", "company", "contract_hours", "hourly_cost"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLoadError(
            f"The employee template is missing required columns: {missing}. "
            f"Found: {list(df.columns)}"
        )

    df["id"] = pd.to_numeric(df["id"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["id"])
    df["id"] = df["id"].astype(int)
    df["contract_hours"] = pd.to_numeric(df["contract_hours"], errors="coerce").fillna(0).astype(float)
    df["hourly_cost"] = pd.to_numeric(df["hourly_cost"], errors="coerce")

    df["company"] = df["company"].apply(
        lambda v: "CC" if ("cc" in str(v).lower() or "combicoach" in str(v).lower()) else "Combibrug"
    )

    if "contract_type" not in df.columns:
        df["contract_type"] = "unknown"
    df["contract_type"] = df["contract_type"].astype(str).str.strip().str.lower()

    df["is_freelancer"] = df["contract_type"].isin(FREELANCER_CONTRACT_TYPES)

    if "end_week" not in df.columns:
        df["end_week"] = pd.NA
    else:
        df["end_week"] = pd.to_numeric(df["end_week"], errors="coerce").astype("Int64")

    if "is_dreammaker" not in df.columns:
        df["is_dreammaker"] = False
    else:
        df["is_dreammaker"] = df["is_dreammaker"].apply(
            lambda v: str(v).strip().lower() in ("true", "1", "yes", "ja", "x")
        )

    df = df[df["contract_hours"] > 0].drop_duplicates(subset=["id"]).reset_index(drop=True)

    missing_rate_ids = df.loc[df["hourly_cost"].isna(), "id"].tolist()
    if missing_rate_ids:
        raise DataLoadError(
            "The employee template has a blank hourly_cost for employee "
            f"id(s) {missing_rate_ids}. hourly_cost is required for every "
            "row, including freelancers — there is no fallback rate. "
            "Please fill it in and re-upload."
        )

    out_cols = ["id", "role", "company", "contract_hours", "hourly_cost",
                "contract_type", "is_freelancer", "end_week", "is_dreammaker"]
    return df[out_cols]

