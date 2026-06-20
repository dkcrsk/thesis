"""
data_loader.py
--------------
Reads the two Combibrug Excel/CSV templates and returns two cleaned
DataFrames:

    - employees_df: one row per employee (permanent or freelancer)
    - locations_df: one row per project location with weekly demand

Only the TEMPLATE format is supported. The historical wide-format files
(medewerkers_cb_cc.xlsx, Kopie van locaties 2025.xlsx) are no longer parsed
directly — they were a major source of crashes (merged cells, missing
sheets, inconsistent date formats) and are not needed now that templates
exist for both employees and demand.

The MILP and the UI never touch raw files directly — they only consume the
two cleaned DataFrames returned here.
"""

import pandas as pd


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ALL_ROLES = [
    "Combicoach", "Stagiair", "Jongerenbegeleider", "Buurtsportcoach",
    "Manager", "Cultuurcoordinator", "Leefstijlcoach", "Other",
]

VALID_COMPANIES = {"Combibrug", "CC"}
VALID_CONTRACT_TYPES = {"permanent", "fixed_term", "oproep", "stage", "freelance"}

_TRUE_STRINGS = {"1", "true", "yes", "ja", "y"}
_FALSE_STRINGS = {"0", "false", "no", "nee", "n"}


# ---------------------------------------------------------------------------
# Custom exception so the UI can show clean messages
# ---------------------------------------------------------------------------

class DataLoadError(Exception):
    """Raised when an input file cannot be parsed.

    Use a clear, planner-friendly message — these are surfaced directly
    in the Streamlit UI.
    """
    pass


# ---------------------------------------------------------------------------
# Small shared helpers
# ---------------------------------------------------------------------------

def _read_table(path_or_buffer, label):
    """Read a .xlsx or .csv file into a DataFrame, with a clear error
    message if anything goes wrong."""
    try:
        name = getattr(path_or_buffer, "name", str(path_or_buffer))
        if str(name).lower().endswith(".csv"):
            df = pd.read_csv(path_or_buffer)
        else:
            df = pd.read_excel(path_or_buffer)
    except Exception as e:
        raise DataLoadError(f"Could not read the {label} file: {e}")

    if df.empty:
        raise DataLoadError(f"The {label} file has no rows.")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    return df


def _check_required_columns(df, required, label):
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLoadError(
            f"The {label} file is missing required columns: {missing}. "
            f"Columns found: {list(df.columns)}."
        )


def _to_bool_flag(series, default_true=True):
    """Convert a column of mixed truthy/falsy values to a clean bool Series.

    Accepts 1/0, True/False, 'yes'/'no', 'ja'/'nee', etc. Blank/NaN values
    fall back to `default_true`.
    """
    def _conv(v):
        if pd.isna(v) or str(v).strip() == "":
            return default_true
        s = str(v).strip().lower()
        if s in _TRUE_STRINGS:
            return True
        if s in _FALSE_STRINGS:
            return False
        return default_true
    return series.apply(_conv)


# ---------------------------------------------------------------------------
# Employee template
# ---------------------------------------------------------------------------

EMPLOYEE_TEMPLATE_COLUMNS = {
    # required
    "id":             "Unique employee number (integer)",
    "role":           "One of: Combicoach, Stagiair, Jongerenbegeleider, "
                      "Buurtsportcoach, Manager, Cultuurcoordinator, "
                      "Leefstijlcoach, Other",
    "company":        "'Combibrug' or 'CC'",
    "contract_type":  "permanent / fixed_term / oproep / stage / freelance",
    "contract_hours": "Weekly hours (decimal). For freelancers this is the "
                      "MAXIMUM they can work this week.",
    # optional — sensible defaults applied when missing
    "is_freelancer":  "1 if freelancer, 0 otherwise. Defaults from contract_type.",
    "available":      "1 if available this week, 0 if on vacation/leave (default 1)",
    "hourly_rate":    "Optional hourly cost override (EUR/h). If left blank, "
                      "the fallback/stage rate from the sidebar is used.",
    "name":           "Display name — not used by the model",
    "notes":          "Free-text notes — not used by the model",
}


def load_employees(path_or_buffer):
    """Load the employee template and return a cleaned employees DataFrame.

    Required columns: id, role, company, contract_type, contract_hours
    Optional columns: is_freelancer, available, hourly_rate, name, notes

    Returns a DataFrame with columns:
        id, role, company, contract_hours, contract_type,
        is_freelancer, available, [hourly_rate]
    """
    df = _read_table(path_or_buffer, "employees")

    required = ["id", "role", "company", "contract_type", "contract_hours"]
    _check_required_columns(df, required, "employees")

    # --- id: must be numeric, drop rows without one
    df["id"] = pd.to_numeric(df["id"], errors="coerce")
    df = df.dropna(subset=["id"])
    if df.empty:
        raise DataLoadError("No valid employee rows after parsing 'id'.")
    df["id"] = df["id"].astype(int)

    # --- contract_hours: must be numeric and positive
    df["contract_hours"] = pd.to_numeric(df["contract_hours"], errors="coerce").fillna(0)

    # --- role: must match one of the canonical roles, else "Other"
    df["role"] = df["role"].astype(str).str.strip()
    df.loc[~df["role"].isin(ALL_ROLES), "role"] = "Other"

    # --- company: normalise free text to 'Combibrug' or 'CC'
    def _norm_company(v):
        s = str(v).strip().lower()
        return "CC" if ("cc" in s or "combicoach" in s) else "Combibrug"
    df["company"] = df["company"].apply(_norm_company)

    # --- contract_type: normalise, fall back to 'fixed_term' if unrecognised
    df["contract_type"] = df["contract_type"].astype(str).str.strip().str.lower()
    df.loc[~df["contract_type"].isin(VALID_CONTRACT_TYPES), "contract_type"] = "fixed_term"

    # --- is_freelancer: explicit column if given, else derived from contract_type
    if "is_freelancer" in df.columns:
        df["is_freelancer"] = _to_bool_flag(df["is_freelancer"], default_true=False)
    else:
        df["is_freelancer"] = df["contract_type"].eq("freelance")
    # Keep the two signals consistent: a freelance contract_type always
    # implies is_freelancer, regardless of what the column said.
    df.loc[df["contract_type"].eq("freelance"), "is_freelancer"] = True

    # --- available: defaults to True
    if "available" in df.columns:
        df["available"] = _to_bool_flag(df["available"], default_true=True)
    else:
        df["available"] = True

    # --- hourly_rate: optional numeric override
    if "hourly_rate" in df.columns:
        df["hourly_rate"] = pd.to_numeric(df["hourly_rate"], errors="coerce")

    # --- drop employees with non-positive contract hours (nothing to assign)
    df = df[df["contract_hours"] > 0].copy()
    if df.empty:
        raise DataLoadError(
            "No employees left after filtering — check that 'contract_hours' "
            "contains positive numbers."
        )

    df = df.drop_duplicates(subset=["id"]).reset_index(drop=True)

    out_cols = ["id", "role", "company", "contract_hours",
                "contract_type", "is_freelancer", "available"]
    if "hourly_rate" in df.columns:
        out_cols.append("hourly_rate")

    return df[out_cols]


def generate_employee_template_example(path):
    """Write a filled-in example employee template to ``path``.

    Contains a realistic mix of permanent staff, stagiaires, and two
    freelancers per company, plus one row showing how to mark someone
    unavailable for the week (vacation example).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = [
        # id, name,                role,                 company,     contract_type, contract_hours, is_freelancer, available, hourly_rate, notes
        (1001, "Anna de Jong",      "Combicoach",         "Combibrug", "permanent",   32,  0, 1, 24.50, "Senior — Combiworld lead"),
        (1002, "Bram van Leeuwen",  "Combicoach",         "Combibrug", "permanent",   28,  0, 1, 22.00, ""),
        (1003, "Carla Smit",        "Jongerenbegeleider", "Combibrug", "permanent",   24,  0, 1, 21.50, ""),
        (1004, "Daan Visser",       "Buurtsportcoach",    "Combibrug", "permanent",   36,  0, 1, 23.00, ""),
        (1005, "Eline Mulder",      "Manager",            "Combibrug", "permanent",   32,  0, 1, 38.00, ""),
        (1006, "Femke Bakker",      "Cultuurcoordinator", "Combibrug", "permanent",   20,  0, 1, 26.00, ""),
        (1007, "Gijs Janssen",      "Leefstijlcoach",     "Combibrug", "permanent",   24,  0, 1, 22.00, ""),
        (1008, "Hanna Klein",       "Stagiair",           "Combibrug", "stage",       16,  0, 1,  8.00, "BBL student"),
        (1009, "Ivo Berg",          "Stagiair",           "Combibrug", "stage",       16,  0, 0,  8.00, "ON VACATION this week"),
        (2001, "Joris Peters",      "Combicoach",         "CC",        "permanent",   32,  0, 1, 25.00, ""),
        (2002, "Kim Vermeer",       "Combicoach",         "CC",        "permanent",   32,  0, 1, 24.00, ""),
        (2003, "Linde de Boer",     "Jongerenbegeleider", "CC",        "permanent",   28,  0, 1, 22.50, ""),
        (2004, "Mark Janssen",      "Buurtsportcoach",    "CC",        "permanent",   36,  0, 1, 23.00, ""),
        (2005, "Noor de Wit",       "Manager",            "CC",        "permanent",   32,  0, 1, 38.00, ""),
        # Freelancers — one or two per company, with high capacity
        (9001, "Freelancer NL-1",   "Combicoach",         "Combibrug", "freelance",   40,  1, 1, 45.00, "ZZP — call only if needed"),
        (9002, "Freelancer NL-2",   "Buurtsportcoach",    "Combibrug", "freelance",   40,  1, 1, 50.00, "ZZP — call only if needed"),
        (9003, "Freelancer CC-1",   "Combicoach",         "CC",        "freelance",   40,  1, 1, 45.00, "ZZP — call only if needed"),
        (9004, "Freelancer CC-2",   "Jongerenbegeleider", "CC",        "freelance",   40,  1, 1, 48.00, "ZZP — call only if needed"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee template"

    headers = ["id", "name", "role", "company", "contract_type",
               "contract_hours", "is_freelancer", "available",
               "hourly_rate", "notes"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    instructions = [
        "Unique number",
        "Display name (not used)",
        "Combicoach / Stagiair / Jongerenbegeleider / Buurtsportcoach / "
        "Manager / Cultuurcoordinator / Leefstijlcoach / Other",
        "'Combibrug' or 'CC'",
        "permanent / fixed_term / oproep / stage / freelance",
        "Weekly hours (max for freelancers)",
        "1 = freelancer, 0 = permanent",
        "1 = available this week, 0 = vacation / leave",
        "EUR/h (optional — leave blank to use sidebar fallback)",
        "Free-text notes (not used)",
    ]
    inst_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                            fill_type="solid")
    for col_idx, inst in enumerate(instructions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=inst)
        cell.fill = inst_fill
        cell.font = Font(italic=True, color="1F4E79")

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    widths = [8, 20, 22, 12, 16, 16, 15, 12, 14, 45]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = w

    wb.save(path)


# ---------------------------------------------------------------------------
# Demand template
# ---------------------------------------------------------------------------

DEMAND_TEMPLATE_COLUMNS = {
    # required
    "location":       "Name of the project location",
    "company":        "Which company runs it: 'Combibrug' or 'CC'",
    "weekly_hours":   "Expected hours of work needed per week (decimal)",
    # optional — sensible defaults applied when missing
    "duration_months": "How many months the project runs (default 12)",
    "division":        "Sub-division label, e.g. 'MDT' (default = company name)",
    "notes":           "Free-text notes — not used by the model",
}


def load_locations(path_or_buffer):
    """Load the demand template and return a cleaned locations DataFrame.

    Required columns: location, company, weekly_hours
    Optional columns: duration_months, division, notes,
                      req_<role> for any role in ALL_ROLES

    Returns a DataFrame with columns:
        location, division, company, weekly_hours, duration_months,
        req_<role> for every role in ALL_ROLES
    """
    df = _read_table(path_or_buffer, "demand")

    required = ["location", "company", "weekly_hours"]
    _check_required_columns(df, required, "demand")

    df = df.dropna(subset=["location", "weekly_hours"])
    df["location"] = df["location"].astype(str).str.strip()
    df["weekly_hours"] = pd.to_numeric(df["weekly_hours"], errors="coerce").fillna(0)
    df = df[df["weekly_hours"] > 0].copy()

    if df.empty:
        raise DataLoadError(
            "The demand file has no valid rows. Check that 'weekly_hours' "
            "contains numbers greater than 0."
        )

    # --- duration_months: optional, default 12
    if "duration_months" in df.columns:
        df["duration_months"] = pd.to_numeric(
            df["duration_months"], errors="coerce"
        ).fillna(12).astype(int)
    else:
        df["duration_months"] = 12

    # --- company: normalise free text
    def _norm_company(v):
        s = str(v).strip().lower()
        return "CC" if ("cc" in s or "combicoach" in s) else "Combibrug"
    df["company"] = df["company"].apply(_norm_company)

    # --- division: optional, default = company name
    if "division" in df.columns:
        df["division"] = df["division"].fillna(df["company"])
    else:
        df["division"] = df["company"]

    # --- per-role headcount requirements: req_<role> columns, default 0
    for r in ALL_ROLES:
        col = f"req_{r.lower()}"
        if col in df.columns:
            df[f"req_{r}"] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
        else:
            df[f"req_{r}"] = 0

    df = df.drop_duplicates(subset=["location"]).reset_index(drop=True)

    out_cols = [
        "location", "division", "company", "weekly_hours", "duration_months",
    ] + [f"req_{r}" for r in ALL_ROLES]

    return df[out_cols]


def generate_demand_template_example(path):
    """Write a filled-in example demand template to ``path``.

    Rosa can use this as a starting point for each new planning cycle.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = [
        # location,                  company,      weekly_hours, duration_months, division, notes
        ("Park Frankendael",         "Combibrug",  12,           10,  "Combiworld",   "School year contract — renewed Jan 2026"),
        ("Sportpark Drieburg",       "Combibrug",   8,           10,  "Combiworld",   "School year contract"),
        ("MDT Watergraafsmeer",      "Combibrug",   6,            4,  "MDT",          "MDT subsidy confirmed Q1 2026"),
        ("MDT Diemen",               "Combibrug",   5,            4,  "MDT",          "MDT subsidy confirmed Q1 2026"),
        ("MDT Bos en Lommer",        "Combibrug",   4,            4,  "MDT",          "Under review — hours may change"),
        ("Olympiaplein",             "CC",         18,           12,  "Combibrug CC", "Year-round location"),
        ("Stadionplein",             "CC",         14,           12,  "Combibrug CC", "Year-round location"),
        ("Apollolaan",               "CC",         10,           12,  "Combibrug CC", "Year-round location"),
        ("Mercatorplein",            "CC",         16,           12,  "Combibrug CC", "Year-round location"),
        ("Surinameplein",            "CC",         12,           12,  "Combibrug CC", "Year-round location"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demand template"

    headers = ["location", "company", "weekly_hours",
               "duration_months", "division", "notes"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    instructions = [
        "Project location name",
        "'Combibrug' or 'CC'",
        "Hours/week (decimal)",
        "Months project runs",
        "Sub-division label",
        "Any notes (not used by model)",
    ]
    inst_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                            fill_type="solid")
    for col_idx, inst in enumerate(instructions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=inst)
        cell.fill = inst_fill
        cell.font = Font(italic=True, color="1F4E79")

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    widths = [28, 12, 14, 17, 18, 45]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = w

    wb.save(path)


# ---------------------------------------------------------------------------
# Per-employee hourly rate file (optional — only needed if the employee
# template doesn't already carry an hourly_rate column for everyone)
# ---------------------------------------------------------------------------

RATES_TEMPLATE_COLUMNS = {
    "id":         "Employee id, matches the employees template",
    "rate_base":  "Base hourly wage (EUR/h)",
    "rate_incl":  "Hourly wage including employer costs (EUR/h)",
    "is_stage":   "1 if this employee is a stagiair (optional)",
}


def load_rates(path_or_buffer):
    """Load an optional rates file: id, rate_base, rate_incl, [is_stage].

    Used by apply_hourly_costs() in model.py as a fallback for any employee
    whose template row doesn't have a direct hourly_rate.
    """
    if path_or_buffer is None:
        return None

    df = _read_table(path_or_buffer, "rates")
    required = ["id"]
    _check_required_columns(df, required, "rates")

    df["id"] = pd.to_numeric(df["id"], errors="coerce")
    df = df.dropna(subset=["id"])
    df["id"] = df["id"].astype(int)

    for col in ("rate_base", "rate_incl"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        else:
            df[col] = None

    if "is_stage" in df.columns:
        df["is_stage"] = _to_bool_flag(df["is_stage"], default_true=False)
    else:
        df["is_stage"] = False

    df = df.drop_duplicates(subset=["id"]).reset_index(drop=True)
    return df[["id", "rate_base", "rate_incl", "is_stage"]]


# ---------------------------------------------------------------------------
# Convenience: load everything at once
# ---------------------------------------------------------------------------

def load_all(employees_path, locations_path, rates_path=None):
    """Load employees, locations, and (optionally) rates.

    All three files must be in the template formats described above.
    Returns (employees_df, locations_df, rates_df).
    """
    employees = load_employees(employees_path)
    locations = load_locations(locations_path)
    rates = load_rates(rates_path) if rates_path else None
    return employees, locations, rates
